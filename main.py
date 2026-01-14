#!/usr/bin/env python3
"""
Merged Sanctions Scraper & Converter
Combines creation.py and conversion.py into one flow:
1. Downloads XML and PDF from SanctionsMap
2. Splits entities and creates Excel template
3. Populates Excel with entity details from XML
"""
import sys, os

def resource_path(relative):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.abspath("."), relative)

CHROME_PATH = resource_path("chromium/chrome-win/chrome.exe")

import re
import os
import sys
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import pdfplumber
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import regex
import gender_guesser.detector as gender
import unicodedata

# ================================================================================
# PART 1: CREATION (Download & Setup)
# ================================================================================

SANCTIONS_URL = "https://www.sanctionsmap.eu/#/main/travel/ban"

import sys
from pathlib import Path

if hasattr(sys, "_MEIPASS"):
    BASE_DIR = Path.cwd()        # folder where EXE was launched from
else:
    BASE_DIR = Path(__file__).resolve().parent

parent_dir = BASE_DIR / "data"
parent_dir.mkdir(exist_ok=True)


xml_folder = parent_dir / "xml_files"
xml_chunks_folder = parent_dir / "xml_chunks"
pdf_folder = parent_dir / "pdf"
pdf_text_chunks_folder = parent_dir / "pdf_text_chunks"

for d in (xml_folder, xml_chunks_folder, pdf_folder, pdf_text_chunks_folder):
    d.mkdir(parents=True, exist_ok=True)

xlsx_path = parent_dir / "sanctions_output.xlsx"

CSV_COLUMNS = [
    "FULL_NAME", "CATEGORY", "F_NAME", "M_NAME", "L_NAME", "GENDER", "DOB",
    "ADD_CITY", "ADD_COUNTRY", "STATE", "NATIONALITIES", "ADDRESS",
    "IDENTITY NUMBER", "IDENTITY TYPE", "REF_DATE", "DETAILS", "WEB_LINK",
    "VIOLATION_ID", "SOURCE", "ALIAS", "ASSOCIATES", "MAIN ACTIVITY",
    "CITIZENSHIP INFORMATION", "STATUS", "REM1", "REM2", "REM3", "REMARKS"
]

DEFAULT_WEB_LINK = "https://www.sanctionsmap.eu/#/main/travel/ban"
DEFAULT_SOURCE = "EU TRAVEL BAN"

def download_url_to_file(url, dest_folder, session=None, timeout=60):
    dest_folder = Path(dest_folder)
    dest_folder.mkdir(parents=True, exist_ok=True)

    if session is None:
        session = requests.Session()
    resp = session.get(url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()

    cd = resp.headers.get("Content-Disposition", "")
    filename = None
    if "filename" in cd.lower():
        m = re.search(r"filename\*?=(?:UTF-8''|\"?)([^\";]+)\"?", cd, flags=re.IGNORECASE)
        if m:
            filename = m.group(1).strip().strip('"')

    if not filename:
        filename = url.split("/")[-1].split("?")[0] or "downloaded_file"

    filename = filename.replace("\\", "_").replace("/", "_")
    dest_path = dest_folder / filename
    with open(dest_path, "wb") as f:
        f.write(resp.content)

    return dest_path


def find_and_download_xml(page):
    print("➡️ Navigating to SanctionsMap to find XML link...")
    page.goto(SANCTIONS_URL, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(7000)

    possible_xpaths = [
        "//ul[@class='filter-list']//li//a//a",
        "//a[contains(@href, 'export') and contains(@href, '.xml')]",
        "//a[contains(@href, '/travelbans/file/') and contains(@href, '.xml')]",
        "//a[contains(@href, '/export') and contains(@href, '.xml')]",
    ]

    href = None
    for xp in possible_xpaths:
        try:
            locator = page.locator(xp).first
            if locator:
                h = locator.get_attribute("href")
                if h:
                    href = h
                    break
        except Exception:
            continue

    if not href:
        all_links = page.locator("//a")
        count = all_links.count()
        for i in range(count):
            try:
                a = all_links.nth(i)
                a_href = a.get_attribute("href") or ""
                if ".xml" in a_href.lower():
                    href = a_href
                    break
            except Exception:
                continue

    if not href:
        raise RuntimeError("Could not find XML link on the page.")

    if href.startswith("/"):
        href = "https://www.sanctionsmap.eu" + href

    print(f"📄 Found XML URL: {href}")
    print("⬇️ Downloading XML with original filename...")
    xml_file_path = download_url_to_file(href, xml_folder)
    print(f"✅ XML saved to: {xml_file_path}")
    return xml_file_path


def split_xml_entities(input_xml_path, output_folder):
    print("🔎 Parsing XML and splitting sanctionEntity tags...")
    tree = ET.parse(input_xml_path)
    root = tree.getroot()

    namespace = ""
    if root.tag.startswith("{"):
        namespace = root.tag.split("}")[0] + "}"

    entities = root.findall(f".//{namespace}sanctionEntity")
    total = len(entities)
    print(f"Found {total} <sanctionEntity> elements")

    for old in Path(output_folder).glob("*.xml"):
        try:
            old.unlink()
        except Exception:
            pass

    seq = 1
    for ent in entities:
        filename = f"entity{seq}.xml"
        out_path = Path(output_folder) / filename

        wrapper = ET.Element("root")
        ent_tree = ET.fromstring(ET.tostring(ent, encoding="utf-8"))
        wrapper.append(ent_tree)

        ET.ElementTree(wrapper).write(out_path, encoding="utf-8", xml_declaration=True)
        seq += 1

    return total


def create_xlsx_with_entity_rows(entity_count, xlsx_file_path):
    print(f"📊 Creating Excel with {entity_count} rows (1 per entity)...")

    if entity_count <= 0:
        df = pd.DataFrame(columns=CSV_COLUMNS)
    else:
        rows = [dict.fromkeys(CSV_COLUMNS, "") for _ in range(entity_count)]
        df = pd.DataFrame(rows, columns=CSV_COLUMNS)

    if "WEB_LINK" in df.columns:
        df["WEB_LINK"] = DEFAULT_WEB_LINK
    if "SOURCE" in df.columns:
        df["SOURCE"] = DEFAULT_SOURCE

    df = df[CSV_COLUMNS]
    df.to_excel(xlsx_file_path, index=False)
    print(f"✅ Excel saved to: {xlsx_file_path}")


def find_and_download_pdf(page):
    print("➡️ Navigating to SanctionsMap to find PDF link...")
    page.goto(SANCTIONS_URL, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(7000)

    possible_pdf_xpaths = [
        "//a[contains(text(),'PDF') and contains(@href, '/travelbans/file/')]",
        "//a[contains(@href, '.pdf') and contains(@href, 'travelbans')]",
        "//a[contains(@href, '.pdf')]",
    ]

    href = None
    for xp in possible_pdf_xpaths:
        try:
            locator = page.locator(xp).first
            if locator:
                h = locator.get_attribute("href")
                if h:
                    href = h
                    break
        except Exception:
            continue

    if not href:
        all_links = page.locator("//a")
        count = all_links.count()
        for i in range(count):
            try:
                a_href = all_links.nth(i).get_attribute("href") or ""
                if ".pdf" in a_href.lower():
                    href = a_href
                    break
            except Exception:
                continue

    if not href:
        raise RuntimeError("Could not find PDF link on the page.")

    if href.startswith("/"):
        href = "https://www.sanctionsmap.eu" + href

    print(f"📄 Found PDF URL: {href}")
    print("⬇️ Downloading PDF with original filename...")
    pdf_file_path = download_url_to_file(href, pdf_folder)
    print(f"✅ PDF saved to: {pdf_file_path}")
    return pdf_file_path


def extract_text_from_pdf(pdf_file_path):
    print("🔥 Extracting text from PDF...")
    full_text = ""
    with pdfplumber.open(pdf_file_path) as pdf:
        for p in pdf.pages:
            text = p.extract_text()
            if text:
                full_text += text + "\n"
    return full_text


def split_entities_from_text(text):
    print("✂️ Splitting PDF text into entity chunks...")
    pattern = r"(?=Entity\s+\d+\b)"
    parts = re.split(pattern, text)
    entities = []

    for chunk in parts:
        chunk = chunk.strip()
        if not chunk:
            continue
        if not chunk.lower().startswith("entity"):
            continue

        prog_m = re.search(r"Programme\s*[:\-]\s*([A-Za-z0-9]+)", chunk, flags=re.IGNORECASE)
        programme = prog_m.group(1).upper() if prog_m else "GEN"

        entities.append({
            "programme": programme,
            "text": chunk
        })

    print(f"🧩 Extracted {len(entities)} entity text chunks from PDF.")
    return entities


def save_text_entities(entities_list, output_folder):
    print("💾 Saving entity text chunks to files...")
    outdir = Path(output_folder)
    outdir.mkdir(parents=True, exist_ok=True)

    for old in outdir.glob("*.txt"):
        try:
            old.unlink()
        except Exception:
            pass

    for idx, ent in enumerate(entities_list, start=1):
        programme = ent["programme"] or "GEN"
        safe_prog = re.sub(r"[^A-Za-z0-9]+", "_", programme).strip("_") or "GEN"
        fname = f"{safe_prog}_entity{idx}.txt"
        fpath = outdir / fname
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(ent["text"])
        print(f" - Saved {fpath}")


# ================================================================================
# PART 2: CONVERSION (Populate Excel)
# ================================================================================

def clean_fullname_no_accents_final(s: str) -> str:
    if not s:
        return ""

    nfkd = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in nfkd if not unicodedata.combining(c))

    repl = {
        "\u2018": "'", "\u2019": "'", "\u201B": "'",
        "\u201C": '"', "\u201D": '"',
        "\u2013": "-", "\u2014": "-",
        "\u00A0": " ",
    }
    for bad, good in repl.items():
        s = s.replace(bad, good)

    s = re.sub(r"[^A-Za-z0-9 .,'\-()]", "", s)
    s = re.sub(r"\s+", " ", s).strip()

    return s.title()


def is_latin_name(text):
    def _normalize_for_latin_check(s: str) -> str:
        if not s:
            return ""
        s = s.strip()

        repl = {
            "\u2018": "'", "\u2019": "'", "\u201B": "'",
            "\u201C": '"', "\u201D": '"', "\u201F": '"',
            "\u00A0": " ", "\u202F": " ",
            "\u2013": "-", "\u2014": "-", "\u2010": "-",
            "\u2011": "-", "\u2012": "-",
        }
        for a, b in repl.items():
            s = s.replace(a, b)

        confusables = {
            "\u0406": "I", "\u0456": "i",
            "\u0401": "E", "\u0451": "e",
        }
        for a, b in confusables.items():
            s = s.replace(a, b)

        s = s.replace('"', ' ')
        s = re.sub(r"\s+", " ", s)
        return s

    norm = _normalize_for_latin_check(text)
    return bool(regex.fullmatch(r"[\p{Latin}0-9 .,'\-()]+", norm))


def clean_name(name):
    name = re.sub(r"\s+", " ", name).strip()
    return name.title()


MALE_TITLES = [
    "mullah", "maulavi", "mawlavi", "moulavi", "molvi", "qari", "ustad",
    "imam", "amir", "haji", "hajji", "agha", "khan", "pir", "sardar",
    "sayed", "sayyid", "syed", "janan agha"
]
MALE_NAME_PATTERNS = [
    "gul ahmad", "gul ahmed", "abdul", "mohammad", "mohammed", "rahman",
    "hakim", "hakimi", "ullah", "uddin", "ishakzai", "noorzai", "zai"
]


def is_forced_male(name):
    if not name:
        return False
    n = name.lower()
    for t in MALE_TITLES:
        if t in n:
            return True
    for p in MALE_NAME_PATTERNS:
        if p in n:
            return True
    return False


def norm_keep_accents(s: str) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", s).strip().lower()


def remove_punctuation(s: str) -> str:
    if not s:
        return ""
    s2 = regex.sub(r"[^\p{L}\p{N}\s]", " ", s)
    return re.sub(r"\s+", " ", s2).strip().lower()


def strip_accents(s: str) -> str:
    if not s:
        return ""
    nfkd = unicodedata.normalize('NFKD', s)
    only_ascii = "".join([c for c in nfkd if not unicodedata.combining(c)])
    return re.sub(r"\s+", " ", only_ascii).strip().lower()


def all_variants(s: str):
    k1 = norm_keep_accents(s)
    k2 = remove_punctuation(s)
    k3 = strip_accents(s)
    return k1, k2, k3


def build_pdf_rem2_mapping(chunks_folder):
    mapping = {}
    if not os.path.exists(chunks_folder):
        print("pdf_text_chunks folder not found:", chunks_folder)
        return mapping

    files = [f for f in os.listdir(chunks_folder) if f.lower().endswith(".txt")]

    for fname in files:
        path = os.path.join(chunks_folder, fname)
        try:
            with open(path, "r", encoding="utf-8") as fh:
                txt = fh.read()
        except:
            try:
                with open(path, "r", encoding="latin-1") as fh:
                    txt = fh.read()
            except Exception:
                continue

        txt = txt.replace("\u00A0", " ").replace("\r", "\n")
        lines = [ln.strip() for ln in txt.splitlines()]

        pdf_fullname = None

        for idx, ln in enumerate(lines):
            m = regex.match(r"(?i)Name\/Alias\s*:\s*(.*)", ln)
            if m:
                candidate = m.group(1).strip()
                if not candidate:
                    j = idx + 1
                    while j < len(lines) and not lines[j].strip():
                        j += 1
                    if j < len(lines):
                        candidate = lines[j].strip()
                if candidate:
                    candidate = regex.split(
                        r"(?i)\b(title|function|birth information|birth date|citizenship information|"
                        r"contact information|identity information|address|remark|url|programme)\b\s*[:]",
                        candidate
                    )[0].strip()
                if candidate and is_latin_name(candidate):
                    pdf_fullname = clean_name(candidate)
                    break

        numbers = []
        programme = None
        i = 0
        while i < len(lines):
            line = lines[i]
            if regex.match(r"(?i)^Number\s*:", line):
                rest = regex.sub(r"(?i)^Number\s*:\s*", "", line).strip()
                if rest:
                    numbers.append(rest)
                else:
                    j = i + 1
                    while j < len(lines) and not lines[j].strip():
                        j += 1
                    if j < len(lines):
                        numbers.append(lines[j].strip())
                    i = j
            if programme is None and regex.match(r"(?i)^Programme\s*:", line):
                rest = regex.sub(r"(?i)^Programme\s*:\s*", "", line).strip()
                if rest:
                    programme = rest
                else:
                    j = i + 1
                    while j < len(lines) and not lines[j].strip():
                        j += 1
                    if j < len(lines):
                        programme = lines[j].strip()
                    i = j
            i += 1

        numbers_clean = [re.sub(r"\s+", " ", n).strip() for n in numbers if n and n.strip()]

        prog_clean = None
        if programme and programme.strip():
            parts = [p.strip() for p in programme.split("|") if p.strip()]
            if parts:
                prog_clean = parts[-1]
            else:
                prog_clean = programme.strip()

        parts = []
        if numbers_clean:
            parts.append("Number: " + " / ".join(numbers_clean))
        if prog_clean:
            parts.append("Programme: " + prog_clean)

        rem2_value = "; ".join(parts) if parts else ""

        if pdf_fullname:
            v1, v2, v3 = all_variants(pdf_fullname)
            if v1 and v1 not in mapping:
                mapping[v1] = rem2_value
            if v2 and v2 not in mapping:
                mapping[v2] = rem2_value
            if v3 and v3 not in mapping:
                mapping[v3] = rem2_value

    return mapping


def populate_full_name():
    excel_path = str(xlsx_path)
    xml_folder_str = str(xml_chunks_folder)
    chunks_folder_str = str(pdf_text_chunks_folder)

    print("\n" + "="*60)
    print("STEP 2: POPULATING EXCEL WITH ENTITY DETAILS")
    print("="*60 + "\n")

    pdf_mapping = build_pdf_rem2_mapping(chunks_folder_str)
    wb = load_workbook(excel_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    detector = gender.Detector(case_sensitive=False)

    files = sorted(
        [f for f in os.listdir(xml_folder_str) if f.startswith("entity") and f.endswith(".xml")],
        key=lambda x: int(regex.findall(r"\d+", x)[0])
    )

    print(f"Found {len(files)} XML entities – PDF mapping entries: {len(pdf_mapping)}")

    full_names = []
    rem2_candidates = []
    current_row = 2

    for file in files:
        xml_path = os.path.join(xml_folder_str, file)
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
        except Exception as e:
            print("Failed parse:", xml_path, e)
            full_names.append("UNKNOWN")
            rem2_candidates.append("")
            ws[f"A{current_row}"].value = "UNKNOWN"
            ws[f"A{current_row}"].fill = yellow_fill
            current_row += 1
            continue

        namespace = ""
        if len(root) > 0 and isinstance(root[0].tag, str) and root[0].tag.startswith("{"):
            namespace = root[0].tag.split("}")[0] + "}"

        # CATEGORY (B)
        category_cell = ws[f"B{current_row}"]
        subject = root.find(f".//{namespace}subjectType")
        if subject is not None:
            classification = subject.attrib.get("classificationCode")
            category_cell.value = classification if classification else "UNKNOWN"
            if not classification:
                category_cell.fill = yellow_fill
        else:
            category_cell.value = "UNKNOWN"
            category_cell.fill = yellow_fill

        # FULL_NAME (A)
        aliases = root.findall(f".//{namespace}nameAlias")
        selected_name = None
        xml_gender_value = None

        for alias in aliases:
            if "gender" in alias.attrib:
                xml_gender_value = alias.attrib["gender"]

            wn = alias.attrib.get("wholeName")
            if wn and is_latin_name(wn):
                selected_name = clean_name(wn)
                break

        full_name_cell = ws[f"A{current_row}"]
        if selected_name:
            full_name_cell.value = selected_name
        else:
            full_name_cell.value = "UNKNOWN"
            full_name_cell.fill = yellow_fill

        full_names.append(selected_name if selected_name else "UNKNOWN")

        # NATIONALITIES (K)
        nat_cell = ws[f"K{current_row}"]
        citizenships = root.findall(f".//{namespace}citizenship")
        if citizenships:
            first_cit = citizenships[0]
            country_desc = first_cit.attrib.get("countryDescription")
            if country_desc and country_desc.strip() and country_desc.strip().upper() != "UNKNOWN":
                nat_cell.value = country_desc.strip().title()
            else:
                nat_cell.value = ""
        else:
            nat_cell.value = ""

        # DOB (G)
        dob_cell = ws[f"G{current_row}"]
        birthdates = root.findall(f".//{namespace}birthdate")
        dob_found = None
        for b in birthdates:
            bd = b.attrib.get("birthdate")
            if bd and bd.strip():
                dob_found = bd.strip()
                break
        if dob_found:
            try:
                yyyy, mm, dd = dob_found.split("-")
                dob_cell.value = f"{dd}-{mm}-{yyyy}"
            except:
                dob_cell.value = ""
        else:
            dob_cell.value = ""

        # ADDRESS – H,I,J
        city_cell = ws[f"H{current_row}"]
        country_cell = ws[f"I{current_row}"]
        state_cell = ws[f"J{current_row}"]

        addresses = root.findall(f".//{namespace}address")

        if addresses:
            first_addr = addresses[0]

            def valid(field):
                return field and field.strip() and field.strip().upper() != "UNKNOWN"

            city_val = first_addr.attrib.get("city")
            if valid(city_val):
                words = city_val.split()
                filtered = []
                i = 0
                while i < len(words):
                    w = words[i]
                    w_clean = re.sub(r"[,.\-;:]", "", w).strip()
                    lw = w_clean.lower()
                    if lw == "province":
                        if filtered:
                            filtered.pop()
                        i += 1
                        continue
                    if lw == "city":
                        i += 1
                        continue
                    if w_clean:
                        filtered.append(w_clean)
                    i += 1

                seen = set()
                unique = []
                for w in filtered:
                    wl = w.lower()
                    if wl not in seen:
                        unique.append(w)
                        seen.add(wl)
                cleaned = " ".join(unique).strip()
                city_cell.value = cleaned if cleaned else ""
            else:
                city_cell.value = ""

            country_val = first_addr.attrib.get("countryDescription")
            if valid(country_val):
                country_cell.value = country_val.strip().title()
            else:
                country_cell.value = ""

            region_val = first_addr.attrib.get("region")
            if valid(region_val):
                words = region_val.split()
                filtered = []
                i = 0
                while i < len(words):
                    w = words[i]
                    w_clean = re.sub(r"[,.\-;:]", "", w).strip()
                    lw = w_clean.lower()
                    if lw == "province":
                        if filtered:
                            filtered.pop()
                        i += 1
                        continue
                    if lw == "city":
                        i += 1
                        continue
                    if w_clean:
                        filtered.append(w_clean)
                    i += 1
                seen = set()
                unique = []
                for w in filtered:
                    wl = w.lower()
                    if wl not in seen:
                        unique.append(w)
                        seen.add(wl)
                cleaned = " ".join(unique).strip()
                state_cell.value = cleaned if cleaned else ""
            else:
                state_cell.value = ""
        else:
            city_cell.value = ""
            country_cell.value = ""
            state_cell.value = ""

        # ADDRESS COLUMN (L)
        addr_cell = ws[f"L{current_row}"]
        address_list = []
        for addr in addresses:
            parts = []
            cd = addr.attrib.get("countryDescription")
            city = addr.attrib.get("city")
            street = addr.attrib.get("street")
            region = addr.attrib.get("region")
            place = addr.attrib.get("place")
            zipcode = addr.attrib.get("zipCode")

            def valid(field):
                return field and field.strip() and field.strip().upper() != "UNKNOWN"

            if valid(cd):
                cd_clean = re.sub(r"\s+", " ", cd.replace(",", " ")).strip()
                parts.append(cd_clean.title())
            for field in [city, street, region, place]:
                if valid(field):
                    cleaned = re.sub(r"\s+", " ", field.replace(",", " ")).strip()
                    parts.append(cleaned)
            if valid(zipcode):
                zip_clean = re.sub(r"\s+", " ", zipcode.replace(",", " ")).strip()
                parts.append(zip_clean)

            if parts:
                address_list.append(" ".join(parts))

        addr_cell.value = "; ".join(address_list) if address_list else ""

        # ALIASES (T)
        alias_cell = ws[f"T{current_row}"]
        all_aliases = []
        selected_latin = selected_name.lower() if selected_name else None

        for alias in aliases:
            wn = alias.attrib.get("wholeName")
            if not wn:
                continue
            if selected_latin and wn.strip().lower() == selected_latin:
                continue
            if is_latin_name(wn):
                all_aliases.append(clean_name(wn))

        alias_cell.value = "; ".join(all_aliases) if all_aliases else ""

        # GENDER (F)
        gender_cell = ws[f"F{current_row}"]
        if xml_gender_value:
            final_gender = "Female" if xml_gender_value.upper() == "F" else "Male"
        else:
            if selected_name and is_forced_male(selected_name):
                final_gender = "Male"
            else:
                if selected_name:
                    first_name = selected_name.split()[0]
                    g = detector.get_gender(first_name)
                    final_gender = "Female" if g == "female" else "Male"
                else:
                    final_gender = "Male"
        gender_cell.value = final_gender

        # REM1 (Y)
        rem1_cell = ws[f"Y{current_row}"]
        all_functions = []
        for alias in aliases:
            func = alias.attrib.get("function")
            if not func:
                continue
            fn = func.strip()
            if re.search(r"\([a-z]\)", fn):
                cleaned = re.sub(r"\([a-z]\)", "|", fn)
                parts = [p.strip().strip(",") for p in cleaned.split("|") if p.strip()]
                all_functions.extend(parts)
            else:
                all_functions.append(fn)

        if all_functions:
            rem1_cell.value = "Designation: " + "; ".join(all_functions)
        else:
            rem1_cell.value = ""

        # REM2 candidate
        rem2_value = ""
        xml_alias_candidates = []

        for alias in aliases:
            wn = alias.attrib.get("wholeName")
            if wn and is_latin_name(wn):
                xml_alias_candidates.append(clean_name(wn))

        if selected_name and selected_name not in xml_alias_candidates:
            xml_alias_candidates.insert(0, selected_name)

        found = False
        for candidate in xml_alias_candidates:
            v1, v2, v3 = all_variants(candidate)
            for key in (v1, v2, v3):
                if key and key in pdf_mapping:
                    rem2_value = pdf_mapping[key]
                    found = True
                    break
            if found:
                break

        # DETAILS COLUMN (P)
        details = {
            "Title": [],
            "Birth date": [],
            "Birth place": [],
            "Citizenship": [],
            "Remark": []
        }

        for reg in root.findall(f".//{namespace}regulation"):
            num_title = reg.attrib.get("numberTitle")
            if num_title:
                details["Title"].append(num_title.strip())

        for alias in root.findall(f".//{namespace}nameAlias"):
            t = alias.attrib.get("title")
            if t:
                cleaned = re.sub(r"\(\w\)", "", t)
                parts = [p.strip() for p in cleaned.split(",") if p.strip()]
                details["Title"].extend(parts)

        birthdates = root.findall(f".//{namespace}birthdate")
        full_date_count = 0
        years_from_full_dates = set()

        for b in birthdates:
            bd = b.attrib.get("birthdate")
            if bd:
                full_date_count += 1
                if full_date_count > 1:
                    try:
                        yyyy, mm, dd = bd.split("-")
                        details["Birth date"].append(f"{dd}-{mm}-{yyyy}")
                        years_from_full_dates.add(yyyy)
                    except:
                        pass
                else:
                    try:
                        yyyy, mm, dd = bd.split("-")
                        years_from_full_dates.add(yyyy)
                    except:
                        pass

        for b in birthdates:
            y = b.attrib.get("year")
            if y and y.isdigit() and y not in years_from_full_dates:
                details["Birth date"].append(y)

        for b in birthdates:
            yr_from = b.attrib.get("yearRangeFrom")
            yr_to = b.attrib.get("yearRangeTo")
            if yr_from and yr_to:
                details["Birth date"].append(f"{yr_from} to {yr_to}")

        for b in birthdates:
            place = b.attrib.get("place")
            if place:
                details["Birth place"].append(place.strip())

        cit_list = []
        for c in root.findall(f".//{namespace}citizenship"):
            d = c.attrib.get("countryDescription")
            if d and d.strip() and d.strip().upper() != "UNKNOWN":
                cit_list.append(d.strip().title())

        if len(cit_list) > 1:
            first = cit_list[0].strip().lower()
            second = cit_list[1].strip()
            if second and second.strip().lower() != first:
                details["Citizenship"] = [second]
            else:
                details["Citizenship"] = []
        else:
            details["Citizenship"] = []

        def clean_remark_text(txt):
            if not txt:
                return None
            t = txt.strip()
            return t if t else None

        for r in root.findall(f".//{namespace}remark"):
            if r.text:
                cleaned = clean_remark_text(r.text)
                if cleaned and cleaned.strip().lower() != "none":
                    details["Remark"].append(cleaned)

        for key in details:
            seen = set()
            uniq = []
            for val in details[key]:
                low = val.lower()
                if low not in seen:
                    seen.add(low)
                    uniq.append(val)
            details[key] = uniq

        parts = []
        order = ["Title", "Birth date", "Birth place", "Citizenship", "Remark"]
        for field in order:
            vals = details[field]
            if not vals:
                continue
            if len(vals) == 1:
                block = f"{field}: {vals[0].strip()}"
            else:
                merged = " / ".join(v.strip() for v in vals)
                block = f"{field}: {merged.strip()}"

            parts.append(block.strip())

        details_value = "; ".join(parts)
        details_value = details_value.replace("\n", " ").replace("\r", " ").strip()
        ws[f"P{current_row}"].value = details_value

        rem2_candidates.append(rem2_value if rem2_value else "")
        current_row += 1

    # SECOND PASS: duplicate-handling for REM2
    total = len(full_names)
    for idx in range(total):
        row = 2 + idx
        fn = full_names[idx]
        cand = rem2_candidates[idx]
        rem2_cell = ws[f"Z{row}"]

        occurrences = [i for i, x in enumerate(full_names) if x == fn]

        if fn == "UNKNOWN":
            rem2_cell.value = ""
            rem2_cell.fill = yellow_fill
            continue

        if len(occurrences) == 1:
            if cand:
                rem2_cell.value = cand
            else:
                rem2_cell.value = ""
                rem2_cell.fill = yellow_fill
        else:
            prev_nonempty = ""
            j = idx - 1
            while j >= 0:
                if rem2_candidates[j]:
                    prev_nonempty = rem2_candidates[j]
                    break
                j -= 1

            next_nonempty = ""
            j = idx + 1
            while j < total:
                if rem2_candidates[j]:
                    next_nonempty = rem2_candidates[j]
                    break
                j += 1

            if prev_nonempty and next_nonempty and prev_nonempty == next_nonempty:
                rem2_cell.value = prev_nonempty
                rem2_candidates[idx] = prev_nonempty
            else:
                rem2_cell.value = ""
                rem2_cell.fill = red_fill

    # THIRD PASS
    for idx in range(total):
        row = 2 + idx
        fn = full_names[idx]
        rem2_cell = ws[f"Z{row}"]

        if fn == "UNKNOWN" or rem2_cell.value:
            continue

        occurrences = [i for i, x in enumerate(full_names) if x == fn]
        if len(occurrences) <= 1:
            continue

        prev_nonempty_cell = ""
        j = idx - 1
        while j >= 0:
            prev_cell_val = ws[f"Z{2 + j}"].value
            if prev_cell_val:
                prev_nonempty_cell = prev_cell_val
                break
            j -= 1

        next_nonempty_cell = ""
        j = idx + 1
        while j < total:
            next_cell_val = ws[f"Z{2 + j}"].value
            if next_cell_val:
                next_nonempty_cell = next_cell_val
                break
            j += 1

        if prev_nonempty_cell and next_nonempty_cell and prev_nonempty_cell == next_nonempty_cell:
            rem2_cell.value = prev_nonempty_cell
            rem2_cell.fill = PatternFill(fill_type=None)

    # FINAL PASS: Color whole row red if column Z is red
    for idx in range(total):
        row = 2 + idx
        rem2_cell = ws[f"Z{row}"]

        fill = rem2_cell.fill
        if fill and fill.start_color:
            rgb = (fill.start_color.rgb or "").upper()
            if rgb.endswith("FF0000"):
                for col in range(2, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill

    # CLEAN FULL_NAME COLUMN (A)
    for row in range(2, ws.max_row + 1):
        cell = ws[f"A{row}"]
        if cell.value and cell.value != "UNKNOWN":
            cell.value = clean_fullname_no_accents_final(cell.value)

    wb.save(excel_path)
    print("\n✅ Excel update complete →", excel_path)


# ================================================================================
# MAIN EXECUTION
# ================================================================================

def run_all():
    try:
        print("\n" + "="*60)
        print("SANCTIONS SCRAPER & CONVERTER - MERGED VERSION")
        print("="*60 + "\n")
        
        print("STEP 1: DOWNLOADING & SETTING UP")
        print("-"*60 + "\n")
        
        with sync_playwright() as pw:
            print("▶ Launching browser (Playwright)...")
            browser = pw.chromium.launch(
            executable_path=CHROME_PATH,
            headless=True,
            slow_mo=120
            )


            context = browser.new_context()
            page = context.new_page()

            # Download XML
            xml_file = None
            try:
                xml_file = find_and_download_xml(page)
            except Exception as e:
                print("⚠️ Warning: XML download failed:", str(e))

            # Download PDF
            pdf_file = None
            try:
                pdf_file = find_and_download_pdf(page)
            except Exception as e:
                print("⚠️ Warning: PDF download failed:", str(e))

            try:
                browser.close()
            except Exception:
                pass

        # Split XML into entity chunks
        entity_count = 0
        if xml_file and Path(xml_file).exists():
            try:
                entity_count = split_xml_entities(xml_file, xml_chunks_folder)
            except Exception as e:
                print("❌ Error while splitting XML entities:", str(e))
                entity_count = 0
        else:
            print("⚠️ No XML file to split.")

        # Create Excel template
        create_xlsx_with_entity_rows(entity_count, xlsx_path)

        # Process PDF text
        if pdf_file and Path(pdf_file).exists():
            try:
                pdf_text = extract_text_from_pdf(pdf_file)
                entities = split_entities_from_text(pdf_text)
                save_text_entities(entities, pdf_text_chunks_folder)
            except Exception as e:
                print("❌ Error processing PDF:", str(e))
        else:
            print("⚠️ No PDF file to process.")

        print("\n" + "="*60)
        print("STEP 1 COMPLETE - FILES CREATED")
        print("="*60)
        print(f"Main folder: {parent_dir}")
        print(f"- XML file: {xml_file if xml_file else 'none'}")
        print(f"- XML chunks: {xml_chunks_folder}")
        print(f"- Excel template: {xlsx_path}")
        print(f"- PDF: {pdf_file if pdf_file else 'none'}")
        print(f"- PDF text chunks: {pdf_text_chunks_folder}")

        # Now run the conversion
        if entity_count > 0:
            populate_full_name()
        else:
            print("\n⚠️ No entities found, skipping conversion step.")

        print("\n" + "="*60)
        print("🎉 ALL DONE - EXCEL FILE POPULATED")
        print("="*60)
        print(f"\nFinal output: {xlsx_path}")

    except Exception as e:
        print("Fatal error:", str(e))
        sys.exit(1)


if __name__ == "__main__":
    run_all()